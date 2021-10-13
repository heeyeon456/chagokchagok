import openpyxl

# 새로운 excel 객체 만들기 , 기본 sheet 만들어짐
def excelWrite():
    wb = openpyxl.Workbook()
    sh = wb.active
    sh.append([10,20,30])
    sh.append([40,50,60])
    sh.append([70,80,90])
    sh.append([30,40,50])
    wb.save('my1.xlsx')

wb = openpyxl.load_workbook('my1.xlsx')
sh = wb['Sheet']
print('row 개수', sh.max_row)
print('column 갯수', sh.max_column)

for c1, c2, c3 in sh:
    print(c1.value, c2.value, c3.value)
print("======================")
for n in sh[1]:
    print(n.value)
print("======================")
for n in sh['A']:
    print(n.value)
print("======================")
for c1, c2 in sh['B2':'C4']:
    print(c1.value, c2.value)
