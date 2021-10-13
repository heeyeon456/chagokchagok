import openpyxl
from openpyxl.chart import Reference, BarChart

class StudentScore:
    def __init__(self, filename):
        self.filename = filename
        #self.student_list = []
        self.num_of_stu = 0

    @classmethod
    def defaultFn(cls):
        print("1, 2, 3, 4 번 중 하나의 숫자를 입력하세요")

    def getMenu(self):
        print("메인 메뉴)\n\t1. 입력\n\t2. 출력\n\t3. 차트\n\t4. 종료")
        try:
            number = int(input("번호를 입력하세요: "))
            menu = {1:self.getInput,
                    2:self.printOut,
                    3:self.print_chart,
                    4:self.exit}
            menu.get(number, self.defaultFn)()
        except:
            print("숫자를 입력하세요")


    def getInput(self):
        # 1. 입력
        isContinue = True
        wb = openpyxl.Workbook()
        sh = wb.active
        sh.append(['이름', '국어', '영어',
                   '수학', '총합', '평균'])
        while isContinue:
            name = input("이름: ")
            korean = int(input("국어: "))
            english = int(input("영어: "))
            math = int(input("수학: "))
            sum = korean + english + math

            sh.append([name, korean, english,
                        math, sum, "%.2f" %(sum / 3)])
            self.num_of_stu += 1
            #self.student_list.append(
            #    {'name': name, 'korean': korean,
            #     'english' :english, 'math': math,
            #     'sum' : sum, 'avg': "%.2f" % (sum / 3)}
            #)
            while 1:
                tmp = input("게속 입력하시겠습니까?(y/n) ")
                print(tmp)
                if tmp == 'n' or tmp == 'y': break
                print("y, n 중 하나의 정답을 입력흐세요. ")
            isContinue = True if tmp == 'y' else False

        wb.save(self.filename)
        self.getMenu()

    def _calculate_stat(self, key, data):
        total_score = 0
        mx = -1
        for d in data:
            total_score += key(d)
            if key(d) > mx :
                mx = key(d)
        avg_score = "%.2f" % (total_score / len(data))

        return total_score, avg_score, mx

    def _title(self):
        print("-"*30)
        print("이름\t국어\t영어\t수학\t총점\t평균")
        print("-"*30)

    def printOut(self):
        # 2. 출력
        total_score_list = []
        self.title()
        wb = openpyxl.load_workbook(self.filename)
        sh = wb['Sheet']
        for s in sh:
            for k in s:
                print("{}".format(k.value), end='\t')
            print()
        print("-"*30)

        result_dict = {}
        for k in list(self.student_list[0].keys())[1:4]:
            result_dict[k] = self._calculate_stat(lambda v: v[k], self.student_list)

        stat_list = ["총점", "평균", "최고점수"]
        for i in range(3):
            print(stat_list[i], end=': ')
            for k in result_dict.keys():
                print(f'{k} {result_dict[k][i]}', end='\t')
            print()
        print()
        self.getMenu()

    def print_chart(self):

        wb = openpyxl.load_workbook(self.filename)
        sh = wb['Sheet']
        chart = BarChart()
        cat = Reference(sh, min_col=1, min_row=2, max_row=5)
        cData = Reference(sh, min_col=2, max_col=5,
                          min_row=1, max_row=self.num_of_stu+1)

        chart.add_data(cData, titles_from_data=True)
        chart.set_categories(cat)
        sh.add_chart(chart, 'H1')
        wb.save(self.filename)

    def __del__(self):
        pass

    def exit(self):
        # 4. 종료
        self.__del__()

if __name__ == "__main__":
    obj = StudentScore("student1.xlsx")
    obj.print_chart()
